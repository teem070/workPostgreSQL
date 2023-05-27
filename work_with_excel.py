from openpyxl import load_workbook
import re

# classroom INTEGER NOT NULL, date DATE NOT NULL, num_pair INTEGER NOT NULL, predmet VARCHAR(200), teacher VARCHAR(200)

path = "C:/Users/Artem/Desktop"
file = "430.xlsx"

file_name = path+'/'+file
print(file_name)

# Load in the workbook
wb = load_workbook(file_name)
sheet = wb.active

# print(sheet.cell(row=1, column=1).value)
# print(sheet.max_column)
# print(sheet.max_row)

max_column = sheet.max_column
max_row = sheet.max_row

day_week = {'пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс'}
classroom = "430"
num_pair = 1

for temp_column in range(1, max_column+1):
    for temp_row in range(1, max_row+1):
        temp_value = sheet.cell(row=temp_row, column=temp_column).value
        if temp_value == None:
            continue
        # if temp_value in day_week:
        #     print("ok")
        #print(temp_value)

        if re.fullmatch(r'...\d{2}.\d{2}.\d{2}', temp_value):
            # print(temp_value)
            match = re.search('\d{2}.\d{2}.\d{2}', temp_value)
            date = match[0]
            num_pair = temp_row
            continue

        if re.search(r'Группа', temp_value):
            # temp_value = sheet.cell(row=temp_row+1, column=temp_column).value
            # if temp_value == None:
            #     continue
            # temp_row += 3
            match_end = re.search('Дисциплина:', temp_value)
            match_start = re.search('Занятие:', temp_value)
            predmet = temp_value[match_end.end()+1:match_start.start()-1]
            match_end = re.search('Преподаватель:', temp_value)
            teacher = temp_value[match_end.end() + 1:]

            print('430', date, temp_row-num_pair, predmet, teacher)

        # else:
        #     print("next")
